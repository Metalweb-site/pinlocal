from __future__ import annotations

from pathlib import Path
import re

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OUTPUT_DIR = Path(__file__).resolve().parent / "outputs"
OUTPUT_PATH = OUTPUT_DIR / "pinlocal-launch-pricing.xlsx"


BLUE = "1D4ED8"
NAVY = "0F172A"
TEXT = "1F2937"
MUTED = "6B7280"
PALE_BLUE = "DBEAFE"
VERY_PALE = "F8FAFC"
GREEN = "16A34A"
YELLOW = "FEF3C7"
WHITE = "FFFFFF"
LIGHT_BORDER = "E5E7EB"


THIN = Side(style="thin", color=LIGHT_BORDER)
BOX_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def money_fmt(symbol: str = "$") -> str:
    return f'"{symbol}"#,##0.00'


def apply_base_sheet_style(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"


def set_col_widths(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def style_cell(cell, *, fill=None, font=None, align=None, border=None, number_format=None) -> None:
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format


def build_assumptions_sheet(wb: Workbook) -> tuple[dict[str, int], dict[str, int]]:
    ws = wb.active
    ws.title = "Assumptions"
    apply_base_sheet_style(ws)
    set_col_widths(ws, {"A": 24, "B": 32, "C": 18, "D": 16, "E": 52})

    ws["A1"] = "PinLocal Launch Budget Assumptions"
    ws["A2"] = "Editable inputs for infra, media, and placeholder market-dependent costs."
    ws.merge_cells("A1:E1")
    ws.merge_cells("A2:E2")
    style_cell(
        ws["A1"],
        font=Font(size=18, bold=True, color=NAVY),
        align=Alignment(horizontal="left"),
    )
    style_cell(ws["A2"], font=Font(size=10, color=MUTED))

    headers = ["Category", "Item", "Unit", "USD / Value", "Notes"]
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=idx, value=header)
        style_cell(
            cell,
            fill=BLUE,
            font=Font(color=WHITE, bold=True),
            align=Alignment(horizontal="center"),
            border=BOX_BORDER,
        )

    assumptions = [
        ("FX", "USD to INR", "INR per USD", 84, "Editable exchange-rate assumption."),
        ("Render", "Frontend Starter", "USD / month", 7, "Frontend web service."),
        ("Render", "Backend Starter", "USD / month", 7, "Cheapest workable backend for private testing."),
        ("Render", "Backend Standard", "USD / month", 25, "Recommended for real city launch with chat and media."),
        ("Render", "Backend Pro", "USD / month", 85, "Scale-up option once concurrency and jobs grow."),
        ("Upstash", "Redis Fixed 250MB", "USD / month", 10, "Recommended because free tier already hit request limit."),
        ("Render", "Postgres 256MB", "USD / month", 6, "Low-end managed DB option."),
        ("Render", "Postgres 1GB", "USD / month", 19, "Practical managed DB option for launch."),
        ("Render", "Postgres 4GB", "USD / month", 75, "Growth-stage DB option."),
        ("Cloudflare R2", "Media budget - friends testing", "USD / month", 0, "R2 is often free at this stage."),
        ("Cloudflare R2", "Media budget - soft launch", "USD / month", 2, "Light image/video usage buffer."),
        ("Cloudflare R2", "Media budget - one city launch", "USD / month", 5, "Useful planning placeholder."),
        ("Cloudflare R2", "Media budget - growth mode", "USD / month", 15, "Higher usage placeholder."),
        ("OTP", "OTP budget - friends testing", "USD / month", 0, "Fill after MSG91 plan/credits are finalized."),
        ("OTP", "OTP budget - soft launch", "USD / month", 0, "Editable placeholder."),
        ("OTP", "OTP budget - one city launch", "USD / month", 0, "Editable placeholder."),
        ("OTP", "OTP budget - growth mode", "USD / month", 0, "Editable placeholder."),
        ("Domain", "Domain monthly budget", "USD / month", 1.5, "Planning placeholder, verify with registrar."),
        ("Taxes", "Taxes / GST placeholder", "USD / month", 0, "Keep zero until exact billing country tax is known."),
    ]

    row_map: dict[str, int] = {}
    category_rows: dict[str, int] = {}
    current_row = 5
    for category, item, unit, value, notes in assumptions:
        ws.cell(row=current_row, column=1, value=category)
        ws.cell(row=current_row, column=2, value=item)
        ws.cell(row=current_row, column=3, value=unit)
        ws.cell(row=current_row, column=4, value=value)
        ws.cell(row=current_row, column=5, value=notes)
        for col in range(1, 6):
            cell = ws.cell(row=current_row, column=col)
            style_cell(
                cell,
                fill=VERY_PALE if current_row % 2 else WHITE,
                font=Font(color=TEXT),
                align=Alignment(vertical="center"),
                border=BOX_BORDER,
            )
        if category in {"OTP", "Domain", "Taxes"}:
            style_cell(ws.cell(row=current_row, column=4), fill=YELLOW, border=BOX_BORDER)
        if unit.startswith("USD"):
            ws.cell(row=current_row, column=4).number_format = money_fmt("$")
        else:
            ws.cell(row=current_row, column=4).number_format = "#,##0.00"

        key = item.lower().replace(" - ", "_").replace(" ", "_").replace("/", "_").replace(".", "")
        key = re.sub(r"_+", "_", key).strip("_")
        row_map[key] = current_row
        category_rows.setdefault(category, current_row)
        current_row += 1

    ws["A26"] = "Planning note"
    ws["B26"] = (
        "Totals in this model are strong for infra planning, but OTP and exact domain tax will need one more pass "
        "before launch because those are vendor- and region-dependent."
    )
    ws.merge_cells("B26:E26")
    style_cell(ws["A26"], font=Font(bold=True, color=NAVY))
    style_cell(ws["B26"], font=Font(color=MUTED), align=Alignment(wrap_text=True))

    return row_map, category_rows


def build_scenarios_sheet(wb: Workbook, rows: dict[str, int]) -> None:
    ws = wb.create_sheet("Scenarios")
    apply_base_sheet_style(ws)
    set_col_widths(ws, {"A": 34, "B": 16, "C": 16, "D": 16, "E": 16, "F": 18})

    ws["A1"] = "Launch Scenarios"
    ws["A2"] = "Four planning modes: from testing with friends to broader growth. All totals are formula-driven from the Assumptions tab."
    ws.merge_cells("A1:F1")
    ws.merge_cells("A2:F2")
    style_cell(ws["A1"], font=Font(size=18, bold=True, color=NAVY))
    style_cell(ws["A2"], font=Font(size=10, color=MUTED))

    scenarios = ["Friends Testing", "Soft Launch", "One City Launch", "Growth Mode"]
    ws["A4"] = "Cost line"
    for idx, scenario in enumerate(scenarios, start=2):
        cell = ws.cell(row=4, column=idx, value=scenario)
        style_cell(
            cell,
            fill=BLUE,
            font=Font(color=WHITE, bold=True),
            align=Alignment(horizontal="center"),
            border=BOX_BORDER,
        )
    style_cell(
        ws["A4"],
        fill=BLUE,
        font=Font(color=WHITE, bold=True),
        align=Alignment(horizontal="center"),
        border=BOX_BORDER,
    )

    def ref(name: str) -> str:
        return f"Assumptions!$D${rows[name]}"

    line_items = [
        ("Render frontend Starter", [f"={ref('frontend_starter')}", f"={ref('frontend_starter')}", f"={ref('frontend_starter')}", f"={ref('frontend_starter')}"]),
        ("Render backend Starter", [f"={ref('backend_starter')}", f"={ref('backend_starter')}", 0, 0]),
        ("Render backend Standard", [0, 0, f"={ref('backend_standard')}", 0]),
        ("Render backend Pro", [0, 0, 0, f"={ref('backend_pro')}"]),
        ("Upstash Redis Fixed 250MB", [f"={ref('redis_fixed_250mb')}", f"={ref('redis_fixed_250mb')}", f"={ref('redis_fixed_250mb')}", f"={ref('redis_fixed_250mb')}"]),
        ("Render Postgres 256MB", [0, f"={ref('postgres_256mb')}", 0, 0]),
        ("Render Postgres 1GB", [0, 0, f"={ref('postgres_1gb')}", 0]),
        ("Render Postgres 4GB", [0, 0, 0, f"={ref('postgres_4gb')}"]),
        ("Cloudflare R2 media budget", [
            f"={ref('media_budget_friends_testing')}",
            f"={ref('media_budget_soft_launch')}",
            f"={ref('media_budget_one_city_launch')}",
            f"={ref('media_budget_growth_mode')}",
        ]),
        ("OTP budget (editable)", [
            f"={ref('otp_budget_friends_testing')}",
            f"={ref('otp_budget_soft_launch')}",
            f"={ref('otp_budget_one_city_launch')}",
            f"={ref('otp_budget_growth_mode')}",
        ]),
        ("Domain budget", [
            f"={ref('domain_monthly_budget')}",
            f"={ref('domain_monthly_budget')}",
            f"={ref('domain_monthly_budget')}",
            f"={ref('domain_monthly_budget')}",
        ]),
        ("Taxes / GST placeholder", [
            f"={ref('taxes_gst_placeholder')}",
            f"={ref('taxes_gst_placeholder')}",
            f"={ref('taxes_gst_placeholder')}",
            f"={ref('taxes_gst_placeholder')}",
        ]),
    ]

    start_row = 5
    for r_idx, (label, values) in enumerate(line_items, start=start_row):
        ws.cell(row=r_idx, column=1, value=label)
        style_cell(
            ws.cell(row=r_idx, column=1),
            fill=VERY_PALE if r_idx % 2 else WHITE,
            font=Font(color=TEXT),
            border=BOX_BORDER,
        )
        for c_idx, value in enumerate(values, start=2):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            fill = YELLOW if "OTP" in label or "Domain" in label or "Taxes" in label else (VERY_PALE if r_idx % 2 else WHITE)
            style_cell(cell, fill=fill, font=Font(color=TEXT), border=BOX_BORDER, number_format=money_fmt("$"))

    infra_total_row = start_row + len(line_items)
    variable_total_row = infra_total_row + 1
    grand_total_row = infra_total_row + 2
    inr_total_row = infra_total_row + 3

    ws.cell(row=infra_total_row, column=1, value="Known infra total (USD)")
    ws.cell(row=variable_total_row, column=1, value="Editable / market-dependent total (USD)")
    ws.cell(row=grand_total_row, column=1, value="Projected monthly total (USD)")
    ws.cell(row=inr_total_row, column=1, value="Projected monthly total (INR)")

    for row in [infra_total_row, variable_total_row, grand_total_row, inr_total_row]:
        style_cell(
            ws.cell(row=row, column=1),
            fill=PALE_BLUE,
            font=Font(bold=True, color=NAVY),
            border=BOX_BORDER,
        )
        for col in range(2, 6):
            cell = ws.cell(row=row, column=col)
            style_cell(cell, fill=PALE_BLUE, font=Font(bold=True, color=NAVY), border=BOX_BORDER)

    for col in range(2, 6):
        letter = get_column_letter(col)
        ws.cell(row=infra_total_row, column=col, value=f"=SUM({letter}5:{letter}13)")
        ws.cell(row=variable_total_row, column=col, value=f"=SUM({letter}14:{letter}16)")
        ws.cell(row=grand_total_row, column=col, value=f"={letter}{infra_total_row}+{letter}{variable_total_row}")
        ws.cell(row=inr_total_row, column=col, value=f"={letter}{grand_total_row}*Assumptions!$D${rows['usd_to_inr']}")
        ws.cell(row=infra_total_row, column=col).number_format = money_fmt("$")
        ws.cell(row=variable_total_row, column=col).number_format = money_fmt("$")
        ws.cell(row=grand_total_row, column=col).number_format = money_fmt("$")
        ws.cell(row=inr_total_row, column=col).number_format = money_fmt("Rs ")

    ws["A22"] = "Recommended live setup"
    ws["B22"] = (
        "For a real city launch, the strongest balance is usually Render Frontend Starter + Render Backend Standard + "
        "Upstash Fixed 250MB + Render Postgres 1GB, with OTP finalized separately."
    )
    ws.merge_cells("B22:F22")
    style_cell(ws["A22"], font=Font(bold=True, color=NAVY))
    style_cell(ws["B22"], font=Font(color=MUTED), align=Alignment(wrap_text=True))


def build_summary_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Summary", 0)
    apply_base_sheet_style(ws)
    set_col_widths(ws, {"A": 18, "B": 18, "C": 18, "D": 18, "E": 18, "F": 18})

    ws["A1"] = "PinLocal Launch Budget"
    ws["A2"] = "Monthly cost planning workbook for launch, one-city growth, and scale-up."
    ws.merge_cells("A1:F1")
    ws.merge_cells("A2:F2")
    style_cell(ws["A1"], font=Font(size=20, bold=True, color=NAVY))
    style_cell(ws["A2"], font=Font(size=11, color=MUTED))

    summary_headers = ["Scenario", "Infra USD", "Variable USD", "Total USD", "Total INR", "Launch note"]
    for idx, header in enumerate(summary_headers, start=1):
        cell = ws.cell(row=4, column=idx, value=header)
        style_cell(
            cell,
            fill=BLUE,
            font=Font(color=WHITE, bold=True),
            align=Alignment(horizontal="center"),
            border=BOX_BORDER,
        )

    notes = {
        "Friends Testing": "Good for testing panels with friends before real public launch.",
        "Soft Launch": "Lowest paid stack that still looks launch-shaped.",
        "One City Launch": "Best recommended production-style starting point.",
        "Growth Mode": "Early scale-up planning, not day-one spend.",
    }
    scenario_cols = {
        "Friends Testing": "B",
        "Soft Launch": "C",
        "One City Launch": "D",
        "Growth Mode": "E",
    }
    for idx, (scenario, col_letter) in enumerate(scenario_cols.items(), start=5):
        ws.cell(row=idx, column=1, value=scenario)
        ws.cell(row=idx, column=2, value=f"=Scenarios!{col_letter}17")
        ws.cell(row=idx, column=3, value=f"=Scenarios!{col_letter}18")
        ws.cell(row=idx, column=4, value=f"=Scenarios!{col_letter}19")
        ws.cell(row=idx, column=5, value=f"=Scenarios!{col_letter}20")
        ws.cell(row=idx, column=6, value=notes[scenario])
        for col in range(1, 7):
            cell = ws.cell(row=idx, column=col)
            style_cell(
                cell,
                fill=VERY_PALE if idx % 2 else WHITE,
                font=Font(color=TEXT),
                border=BOX_BORDER,
                align=Alignment(vertical="center", wrap_text=(col == 6)),
            )
        for col in [2, 3, 4]:
            ws.cell(row=idx, column=col).number_format = money_fmt("$")
        ws.cell(row=idx, column=5).number_format = money_fmt("Rs ")

    ws["A11"] = "What this means"
    ws["A12"] = "Current must-pay unlock"
    ws["B12"] = "Upstash Fixed 250MB"
    ws["C12"] = "$10/mo"
    ws["D12"] = "Needed because your free Redis request limit is already exhausted."
    ws.merge_cells("D12:F12")
    ws["A13"] = "Recommended launch stack"
    ws["B13"] = "Frontend Starter + Backend Standard + Upstash Fixed + Postgres 1GB"
    ws["C13"] = "=Scenarios!D19"
    ws["D13"] = "Good balance for launch without overspending too early."
    ws.merge_cells("D13:F13")
    ws["A14"] = "Biggest unknown"
    ws["B14"] = "OTP cost"
    ws["C14"] = "Editable"
    ws["D14"] = "MSG91 pricing should be filled in once your plan or credit pack is finalized."
    ws.merge_cells("D14:F14")
    for row in [11, 12, 13, 14]:
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            style_cell(
                cell,
                fill=PALE_BLUE if row == 11 else WHITE,
                font=Font(bold=(row == 11 or col == 1), color=NAVY if row == 11 else TEXT),
                border=BOX_BORDER,
                align=Alignment(wrap_text=True),
            )
    ws["C13"].number_format = money_fmt("$")

    chart = BarChart()
    chart.title = "Projected Monthly Budget by Scenario (INR)"
    chart.y_axis.title = "Scenario"
    chart.x_axis.title = "INR"
    data = Reference(ws, min_col=5, min_row=4, max_row=8)
    categories = Reference(ws, min_col=1, min_row=5, max_row=8)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 8
    chart.width = 16
    chart.legend = None
    ws.add_chart(chart, "A17")

    ws["E12"].number_format = money_fmt("$")


def build_subscriptions_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Subscriptions")
    apply_base_sheet_style(ws)
    set_col_widths(ws, {"A": 24, "B": 28, "C": 30, "D": 44})

    ws["A1"] = "Subscriptions To Buy"
    ws["A2"] = "Clean list of what you need first, what can wait, and why."
    ws.merge_cells("A1:D1")
    ws.merge_cells("A2:D2")
    style_cell(ws["A1"], font=Font(size=18, bold=True, color=NAVY))
    style_cell(ws["A2"], font=Font(size=10, color=MUTED))

    headers = ["Priority", "Service", "Recommended plan", "Why"]
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=idx, value=header)
        style_cell(
            cell,
            fill=BLUE,
            font=Font(color=WHITE, bold=True),
            align=Alignment(horizontal="center"),
            border=BOX_BORDER,
        )

    rows = [
        ("Buy now", "Upstash Redis", "Fixed 250MB", "Free quota is already exhausted, so Redis is the first paid unlock."),
        ("Buy now", "Render Frontend", "Starter", "Enough for the website frontend at launch."),
        ("Buy now", "Render Backend", "Standard", "Safer for chat, uploads, sockets, OTP, and jobs."),
        ("Buy now / optional", "Managed Postgres", "Render Postgres 1GB", "Use this if you want a cleaner single-vendor launch setup."),
        ("Later", "Cloudflare R2", "Pay only for usage", "Usually tiny spend at the beginning, but it is the right long-term media layer."),
        ("Later", "Backend upgrade", "Render Pro", "Only when one-city traffic starts pushing CPU, workers, or concurrency."),
        ("Later", "Higher DB plan", "Postgres 4GB", "Only when storage, connections, or query load demand it."),
    ]
    for r_idx, row in enumerate(rows, start=5):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            style_cell(
                cell,
                fill=VERY_PALE if r_idx % 2 else WHITE,
                font=Font(color=TEXT),
                border=BOX_BORDER,
                align=Alignment(wrap_text=True),
            )


def build_sources_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Sources")
    apply_base_sheet_style(ws)
    set_col_widths(ws, {"A": 22, "B": 58, "C": 42})

    ws["A1"] = "Pricing Sources"
    ws["A2"] = "Official links used or referenced for this planning workbook."
    ws.merge_cells("A1:C1")
    ws.merge_cells("A2:C2")
    style_cell(ws["A1"], font=Font(size=18, bold=True, color=NAVY))
    style_cell(ws["A2"], font=Font(size=10, color=MUTED))

    headers = ["Vendor", "URL", "Comment"]
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=idx, value=header)
        style_cell(
            cell,
            fill=BLUE,
            font=Font(color=WHITE, bold=True),
            align=Alignment(horizontal="center"),
            border=BOX_BORDER,
        )

    sources = [
        ("Render", "https://render.com/pricing", "Frontend, backend, Postgres pricing references."),
        ("Upstash", "https://upstash.com/pricing", "Redis fixed plan and free-tier request limit context."),
        ("Cloudflare R2", "https://developers.cloudflare.com/r2/pricing/", "Usage-based media storage pricing."),
        ("Supabase", "https://supabase.com/pricing", "Keep as a check if you stay on Supabase instead of moving DB."),
        ("MSG91 OTP", "https://msg91.com/in/otp", "Use dashboard/sales confirmation for final OTP numbers."),
    ]
    for r_idx, row in enumerate(sources, start=5):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            style_cell(
                cell,
                fill=VERY_PALE if r_idx % 2 else WHITE,
                font=Font(color=TEXT),
                border=BOX_BORDER,
                align=Alignment(wrap_text=True),
            )


def build_workbook() -> Path:
    wb = Workbook()
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    row_map, _ = build_assumptions_sheet(wb)
    build_scenarios_sheet(wb, row_map)
    build_summary_sheet(wb)
    build_subscriptions_sheet(wb)
    build_sources_sheet(wb)

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.row not in (1, 2) and cell.column == 1 and not cell.border.left.style:
                    cell.border = BOX_BORDER

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)

    # Sanity-check the artifact so we do not hand over a broken workbook.
    loaded = load_workbook(OUTPUT_PATH, data_only=False)
    assert "Summary" in loaded.sheetnames
    assert loaded["Summary"]["A1"].value == "PinLocal Launch Budget"
    return OUTPUT_PATH


if __name__ == "__main__":
    path = build_workbook()
    print(path)
